from __future__ import annotations

import json
import math
from collections import Counter
from pathlib import Path

from openpyxl import load_workbook


SITE_XLSX = Path("/Users/tt2000/Documents/天安智联/AI/项目管理工具包/Project Data/点位管理表-无锡车路云.xlsx")
DATA_JS = Path("prototype/data.js")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def number(value) -> float:
    raw = text(value).replace(",", "")
    try:
        parsed = float(raw)
    except ValueError:
        return 0
    return parsed if math.isfinite(parsed) else 0


def truthy(value) -> bool:
    return text(value) in {"是", "有", "1", "true", "TRUE", "√", "Y", "y"}


def normalize_district(value) -> str:
    raw = text(value)
    if "锡山" in raw:
        return "锡山"
    if "新吴" in raw or "经开" in raw:
        return "新吴"
    if "惠山" in raw:
        return "惠山"
    if "梁溪" in raw:
        return "梁溪"
    if "滨湖" in raw:
        return "滨湖"
    return raw or "未标注"


def out_of_china(lng: float, lat: float) -> bool:
    return lng < 72.004 or lng > 137.8347 or lat < 0.8293 or lat > 55.8271


def transform_lat(x: float, y: float) -> float:
    ret = -100 + 2 * x + 3 * y + 0.2 * y * y + 0.1 * x * y + 0.2 * math.sqrt(abs(x))
    ret += ((20 * math.sin(6 * x * math.pi) + 20 * math.sin(2 * x * math.pi)) * 2) / 3
    ret += ((20 * math.sin(y * math.pi) + 40 * math.sin((y / 3) * math.pi)) * 2) / 3
    ret += ((160 * math.sin((y / 12) * math.pi) + 320 * math.sin((y * math.pi) / 30)) * 2) / 3
    return ret


def transform_lng(x: float, y: float) -> float:
    ret = 300 + x + 2 * y + 0.1 * x * x + 0.1 * x * y + 0.1 * math.sqrt(abs(x))
    ret += ((20 * math.sin(6 * x * math.pi) + 20 * math.sin(2 * x * math.pi)) * 2) / 3
    ret += ((20 * math.sin(x * math.pi) + 40 * math.sin((x / 3) * math.pi)) * 2) / 3
    ret += ((150 * math.sin((x / 12) * math.pi) + 300 * math.sin((x / 30) * math.pi)) * 2) / 3
    return ret


def wgs84_to_gcj02(lng: float, lat: float) -> tuple[float, float, bool]:
    if not lng or not lat or out_of_china(lng, lat):
        return lng, lat, False
    a = 6378245
    ee = 0.006693421622965943
    d_lat = transform_lat(lng - 105, lat - 35)
    d_lng = transform_lng(lng - 105, lat - 35)
    rad_lat = (lat / 180) * math.pi
    magic = math.sin(rad_lat)
    magic = 1 - ee * magic * magic
    sqrt_magic = math.sqrt(magic)
    d_lat = (d_lat * 180) / (((a * (1 - ee)) / (magic * sqrt_magic)) * math.pi)
    d_lng = (d_lng * 180) / ((a / sqrt_magic) * math.cos(rad_lat) * math.pi)
    return lng + d_lng, lat + d_lat, True


def device_type(device: dict) -> str:
    name = text(device.get("name"))
    material = text(device.get("materialCode"))
    if "RSU" in name or material.startswith("RSU"):
        return "RSU"
    if "雷达" in name:
        return "感知设备"
    if "相机" in name or "摄像" in name:
        return "视频设备"
    if "信号" in name:
        return "信号系统"
    if "抱杆箱" in name or "机柜" in name:
        return "机箱机柜"
    if "标志" in name:
        return "交通标志"
    return "其他设备"


def read_prototype_data() -> dict:
    raw = DATA_JS.read_text(encoding="utf-8")
    prefix = "window.PROTOTYPE_DATA = "
    if not raw.startswith(prefix):
        raise RuntimeError("prototype/data.js does not start with window.PROTOTYPE_DATA assignment")
    return json.loads(raw[len(prefix) :].rstrip().rstrip(";"))


def read_site_rows() -> list[dict]:
    wb = load_workbook(SITE_XLSX, read_only=True, data_only=True, keep_links=False)
    ws = wb["点位管理表"]
    rows = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    data = []
    for values in rows:
        item = {headers[index]: values[index] if index < len(values) else None for index in range(len(headers))}
        data.append(item)
    wb.close()
    return data


def build_site(row: dict, legacy: dict | None) -> dict:
    node_id = text(row.get("NodeID"))
    raw_lng = number(row.get("原始规划经度(CGCS2000)"))
    raw_lat = number(row.get("原始规划纬度(CGCS2000)"))
    lng_gcj, lat_gcj, converted = wgs84_to_gcj02(raw_lng, raw_lat)
    devices = legacy.get("devices", []) if legacy else []
    device_record_count = len(devices)
    device_type_count = len({device_type(device) for device in devices})
    installed_qty = sum(number(device.get("installedQty")) for device in devices)
    issue_count = sum(1 for device in devices if text(device.get("issue"))) + (0 if converted else 1)
    return {
        "serialNo": text(row.get("序号")),
        "type": text(row.get("点位类型")) or "未标注",
        "district": normalize_district(row.get("行政区域")),
        "originalDistrict": text(row.get("行政区域")),
        "nodeId": node_id,
        "crossId": text(row.get("CrossID")),
        "name": text(row.get("路口名称")) or f"未命名点位 {node_id}",
        "englishName": text(row.get("路口英文名称（NodeName）")),
        "vendor": text(row.get("信号机厂商")) or "待确认",
        "perception": text(row.get("感知点位")) or "未标注",
        "adaptive": truthy(row.get("是否为自适应路口")),
        "scope": truthy(row.get("是否计入本期信号系统改造范围")),
        "variableLaneCount": number(row.get("可变车道数量")),
        "variableLaneDirection": text(row.get("可变车道方向")),
        "lngCgcs": raw_lng,
        "latCgcs": raw_lat,
        "lngGcj": lng_gcj or 120.31,
        "latGcj": lat_gcj or 31.49,
        "coordinateConvertMethod": "cgcs2000_as_wgs84_to_gcj02",
        "status": legacy.get("status", "已规划") if legacy else "已规划",
        "deviceCount": legacy.get("deviceCount", device_type_count) if legacy else 0,
        "deviceRecordCount": device_record_count,
        "deviceTypeCount": device_type_count,
        "installedQty": installed_qty,
        "issueCount": issue_count,
        "archiveCompleteness": legacy.get("archiveCompleteness", 30) if legacy else 30,
        "devices": devices,
    }


def sort_key(site: dict):
    node_id = text(site.get("nodeId"))
    return (0, int(node_id)) if node_id.isdigit() else (1, node_id)


def main() -> None:
    data = read_prototype_data()
    legacy_by_node = {text(site.get("nodeId")): site for site in data.get("sites", []) if text(site.get("nodeId"))}
    rows = read_site_rows()
    imported = []
    blank_node_rows = 0
    for row in rows:
        node_id = text(row.get("NodeID"))
        if not node_id:
            blank_node_rows += 1
            continue
        imported.append(build_site(row, legacy_by_node.get(node_id)))
    ids = [site["nodeId"] for site in imported]
    duplicates = [node_id for node_id, count in Counter(ids).items() if count > 1]
    if duplicates:
        raise RuntimeError(f"Duplicate NodeID in imported sites: {duplicates[:20]}")
    imported.sort(key=sort_key)
    data["sites"] = imported
    data.setdefault("stats", {})
    data["stats"]["siteTotal"] = len(imported)
    data["stats"]["prototypeSites"] = len(imported)
    district_order = ["锡山", "惠山", "滨湖", "新吴", "梁溪"]
    present = {site["district"] for site in imported if site["district"] and site["district"] != "未标注"}
    data["stats"]["districts"] = [district for district in district_order if district in present] + sorted(present - set(district_order))
    DATA_JS.write_text("window.PROTOTYPE_DATA = " + json.dumps(data, ensure_ascii=False, indent=2) + ";\n", encoding="utf-8")
    print(f"imported_sites={len(imported)}")
    print(f"blank_nodeid_rows_skipped={blank_node_rows}")
    print(f"legacy_sites_before={len(legacy_by_node)}")
    print(f"preserved_device_sites={sum(1 for site in imported if site['devices'])}")
    print(f"total_devices_preserved={sum(len(site['devices']) for site in imported)}")
    print(f"districts={','.join(data['stats']['districts'])}")


if __name__ == "__main__":
    main()
