from __future__ import annotations

import argparse
import shutil
from collections import Counter
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_BASE = Path("/Users/tt2000/Documents/天安智联/AI/项目管理工具包/Project Data/点位管理表-无锡车路云.bak-20260603-141258.xlsx")
CURRENT_TARGET = Path("/Users/tt2000/Documents/天安智联/AI/项目管理工具包/Project Data/点位管理表-无锡车路云.xlsx")
LATEST_PLAN = Path("/Users/tt2000/Documents/天安智联/AI/项目管理工具包/Project Data/无锡车路云一体化点位NODE规划清单-Node点位规划信息 (1).xlsx")
DEFAULT_OUTPUT = Path("outputs/点位管理表-无锡车路云-保留原表头导入最新Node规划.xlsx")


def text(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def clean_district(value):
    raw = text(value)
    return raw[:-1] if raw.endswith("区") else raw


def header_map(ws):
    return {text(cell.value): cell.column for cell in ws[1]}


def assert_unique(label: str, ids: list[str]) -> None:
    duplicates = [node_id for node_id, count in Counter(ids).items() if count > 1]
    if duplicates:
        raise RuntimeError(f"{label} NodeID duplicates found: {duplicates[:20]}")


def read_latest_plan_rows(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True, keep_links=False)
    ws = wb.worksheets[0]
    rows = ws.iter_rows(values_only=True)
    headers = [text(value) for value in next(rows)]
    data = []
    for values in rows:
        data.append({headers[index]: values[index] if index < len(values) else None for index in range(len(headers))})
    wb.close()
    return headers, data


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    for column in range(1, ws.max_column + 1):
        source = ws.cell(source_row, column)
        target = ws.cell(target_row, column)
        if source.has_style:
            target._style = copy(source._style)
        target.number_format = source.number_format
        target.alignment = copy(source.alignment)
        target.border = copy(source.border)
        target.fill = copy(source.fill)
        target.font = copy(source.font)


def sort_key(node_id: str):
    return (0, int(node_id)) if node_id.isdigit() else (1, node_id)


def build_output(base_path: Path, output_path: Path) -> dict[str, object]:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(base_path, output_path)

    wb = load_workbook(output_path)
    ws = wb["点位管理表"]
    old_headers = header_map(ws)
    latest_headers, latest_rows = read_latest_plan_rows(LATEST_PLAN)

    required_old = [
        "序号",
        "点位类型",
        "行政区域",
        "NodeID",
        "CrossID",
        "路口英文名称（NodeName）",
        "原始规划经度(CGCS2000)",
        "原始规划纬度(CGCS2000)",
        "路口名称",
        "信号机厂商",
        "感知点位",
    ]
    required_latest = [
        "序号",
        "路口类型",
        "行政区",
        "Node ID",
        "CrossID",
        "Node Name",
        "原始规划经度(CGCS2000)",
        "原始规划纬度(CGCS2000)",
        "路口名称",
        "信号机品牌",
        "感知点位类型级别",
    ]
    missing_old = [header for header in required_old if header not in old_headers]
    missing_latest = [header for header in required_latest if header not in latest_headers]
    if missing_old or missing_latest:
        raise RuntimeError(f"Missing headers: old={missing_old}, latest={missing_latest}")

    old_node_rows = {}
    old_ids = []
    for row in range(2, ws.max_row + 1):
        node_id = text(ws.cell(row, old_headers["NodeID"]).value)
        if node_id:
            old_node_rows[node_id] = row
            old_ids.append(node_id)

    latest_by_node = {}
    latest_ids = []
    for row in latest_rows:
        node_id = text(row.get("Node ID"))
        if node_id:
            latest_by_node[node_id] = row
            latest_ids.append(node_id)

    assert_unique("old", old_ids)
    assert_unique("latest", latest_ids)

    # Only update columns that already exist in the original point table.
    # Other original columns are preserved for existing rows and left blank for appended rows.
    mappings = [
        ("序号", lambda r: r.get("序号")),
        ("点位类型", lambda r: text(r.get("路口类型")) or None),
        ("行政区域", lambda r: clean_district(r.get("行政区")) or None),
        ("NodeID", lambda r: r.get("Node ID")),
        ("CrossID", lambda r: r.get("CrossID")),
        ("路口英文名称（NodeName）", lambda r: r.get("Node Name")),
        ("原始规划经度(CGCS2000)", lambda r: r.get("原始规划经度(CGCS2000)")),
        ("原始规划纬度(CGCS2000)", lambda r: r.get("原始规划纬度(CGCS2000)")),
        ("路口名称", lambda r: text(r.get("路口名称")) or None),
        ("信号机厂商", lambda r: text(r.get("信号机品牌")) or None),
        ("感知点位", lambda r: text(r.get("感知点位类型级别")) or None),
    ]

    updated_rows = 0
    changed_cells = 0
    for node_id in sorted(set(old_node_rows) & set(latest_by_node), key=sort_key):
        target_row = old_node_rows[node_id]
        source = latest_by_node[node_id]
        changed = False
        for header, getter in mappings:
            cell = ws.cell(target_row, old_headers[header])
            value = getter(source)
            if text(cell.value) != text(value):
                cell.value = value
                changed = True
                changed_cells += 1
        if changed:
            updated_rows += 1

    appended = 0
    style_source_row = max(2, ws.max_row)
    for node_id in sorted(set(latest_by_node) - set(old_node_rows), key=sort_key):
        source = latest_by_node[node_id]
        target_row = ws.max_row + 1
        copy_row_style(ws, style_source_row, target_row)
        for header, getter in mappings:
            ws.cell(target_row, old_headers[header]).value = getter(source)
        appended += 1

    final_ids = [
        text(ws.cell(row, old_headers["NodeID"]).value)
        for row in range(2, ws.max_row + 1)
        if text(ws.cell(row, old_headers["NodeID"]).value)
    ]
    assert_unique("final", final_ids)
    wb.save(output_path)
    return {
        "output": str(output_path.resolve()),
        "header_count": ws.max_column,
        "old_unique_nodeids": len(set(old_ids)),
        "latest_unique_nodeids": len(set(latest_ids)),
        "intersection": len(set(old_ids) & set(latest_ids)),
        "new_only_appended": appended,
        "old_only_preserved": len(set(old_ids) - set(latest_ids)),
        "updated_rows": updated_rows,
        "changed_cells": changed_cells,
        "final_rows_with_header": ws.max_row,
        "final_unique_nodeids": len(set(final_ids)),
        "final_duplicate_nodeids": 0,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--base", type=Path, default=DEFAULT_BASE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    args = parser.parse_args()
    stats = build_output(args.base, args.output)
    for key, value in stats.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
